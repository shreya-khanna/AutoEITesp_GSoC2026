"""
Merge human transcriptions (Excel) with ASR outputs (data/transcribed/v2, v2 filtered to
word-probability > 0.6 via build_summary_filtered.py, data/transcribed/parakeet, data/transcribed/wav2vec2)
into one CSV per recording, under data/postprocessed/v2/version_a|b/<recording>/.

Recording folders are matched to their human-transcription Excel sheet by (recording number,
version letter) parsed out of the folder / sheet name -- both are written inconsistently across
the dataset (e.g. "038.075_EITv-2A" vs "038075_EIT-2A", "038.075-2A" vs "38.075-2A" vs
"38.075-B"), so both sides are normalized before matching. Rows are aligned by rank position
(1st ASR item <-> Excel row "Sentence" 1, 2nd <-> 2, ...) since the raw ASR item index carries a
recording-specific offset (leading instruction/intro audio) instead of starting at 1.

For "hit" recordings (item_count == 30 in data/transcribed/v2, see
data/transcribed_v2_hit_recordings.txt), each ASR source's per-item text is also scored against
both human raters with WER, MER, and CER (via jiwer), after light normalization (lowercase,
punctuation stripped) so formatting-only differences don't inflate the score. Scoring is
restricted to hit recordings because only there is the rank-based item<->Sentence alignment
described above known to be trustworthy.

Usage:
    python -m src.postprocessing.build_postprocessed_csv
"""
import csv
import json
import logging
import re
from collections import defaultdict
from pathlib import Path

import jiwer
import openpyxl

from src.utils.paths import data_path

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

VERSIONS = ["version_a", "version_b"]
ASR_SOURCES = {
    "v2": ("transcribed/v2", "summary.json"),
    "v2_filtered": ("transcribed/v2", "summary_filtered.json"),
    "parakeet": ("transcribed/parakeet", "summary.json"),
    "wav2vec2": ("transcribed/wav2vec2", "summary.json"),
}
HUMAN_RATERS = ["rater1", "rater2"]

FOLDER_RE = re.compile(r"^038\.?(\d{3})")
SHEET_RE = re.compile(r"^0?38\.(\d{3})-(\d)?([AB])")

HUMAN_COLUMNS = {
    "sentence": "sentence_number",
    "stimulus": "stimulus",
    "transcription rater 1": "human_rater1_text",
    "transcription rater 2": "human_rater2_text",
}


def parse_folder_name(name: str) -> tuple[str, str] | None:
    """Return (recording_number, version_letter) e.g. ("075", "A") from a recording folder name."""
    m = FOLDER_RE.match(name)
    if not m:
        return None
    nnn = m.group(1)
    rest = re.sub(r"\.mp3$", "", name[m.end():])
    letters = re.findall(r"[AB]", rest)
    if not letters:
        return None
    return nnn, letters[-1]


def parse_sheet_name(name: str) -> tuple[str, str] | None:
    """Return (recording_number, version_letter) from an Excel sheet name."""
    m = SHEET_RE.match(name.strip())
    if not m:
        return None
    return m.group(1), m.group(3)


def load_human_sheets(human_dir: Path) -> dict[tuple[str, str], list[dict]]:
    """Map (recording_number, version_letter) -> list of per-Sentence row dicts.

    When the same key appears in more than one sheet (file ranges overlap at their
    boundaries), the sheet with actual rater data wins over an empty/template sheet.
    """
    candidates: dict[tuple[str, str], list[tuple[str, str, list[dict]]]] = defaultdict(list)

    for xlsx_path in sorted(human_dir.glob("*.xlsx")):
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            key = parse_sheet_name(sheet_name)
            if key is None:
                continue
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter)
            col_map = {}
            for i, h in enumerate(header):
                if h is None:
                    continue
                norm = str(h).strip().lower()
                if norm in HUMAN_COLUMNS:
                    col_map[HUMAN_COLUMNS[norm]] = i

            rows = []
            for row in rows_iter:
                if row[0] is None:
                    break
                record = {}
                for out_col, idx in col_map.items():
                    record[out_col] = row[idx] if idx < len(row) else None
                rows.append(record)
            candidates[key].append((xlsx_path.name, sheet_name, rows))
        wb.close()

    result = {}
    for key, versions in candidates.items():
        if len(versions) == 1:
            result[key] = versions[0][2]
            continue
        # Prefer the sheet with the most non-empty rater-1-transcription rows.
        best = max(
            versions,
            key=lambda v: sum(1 for r in v[2] if r.get("human_rater1_text")),
        )
        log.info(
            "Duplicate human sheet for %s: choosing %s/%s over %s",
            key,
            best[0],
            best[1],
            [f"{f}/{s}" for f, s, _ in versions if (f, s) != (best[0], best[1])],
        )
        result[key] = best[2]
    return result


def load_asr_items(recording_dir: Path, summary_filename: str) -> list[dict]:
    summary_path = recording_dir / summary_filename
    if not summary_path.exists():
        return []
    with open(summary_path, encoding="utf-8") as f:
        summary = json.load(f)
    items = sorted(summary.get("items", []), key=lambda it: it["index"])
    return items


def normalize_text(text) -> str:
    if not text:
        return ""
    text = str(text).lower()
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def compute_wer_mer_cer(reference, hypothesis) -> tuple[float | None, float | None, float | None]:
    ref_norm = normalize_text(reference)
    if not ref_norm:
        return None, None, None
    hyp_norm = normalize_text(hypothesis)
    result = jiwer.process_words(ref_norm, hyp_norm)
    cer = jiwer.cer(ref_norm, hyp_norm)
    return round(result.wer, 4), round(result.mer, 4), round(cer, 4)


def build_rows(
    human_rows: list[dict], asr_by_source: dict[str, list[dict]], score_quality: bool
) -> list[dict]:
    max_sentence = max(
        [len(rows) for rows in [human_rows, *asr_by_source.values()]] + [0]
    )

    human_by_num = {r["sentence_number"]: r for r in human_rows if r.get("sentence_number")}
    asr_by_source_ranked = {
        source: {i + 1: item for i, item in enumerate(items)}
        for source, items in asr_by_source.items()
    }

    rows = []
    for n in range(1, max_sentence + 1):
        row = {"sentence_number": n}
        h = human_by_num.get(n, {})
        row["stimulus"] = h.get("stimulus")
        row["human_rater1_text"] = h.get("human_rater1_text")
        row["human_rater2_text"] = h.get("human_rater2_text")
        # Inter-rater agreement: rater1 as reference, rater2 as hypothesis. Unlike ASR scoring,
        # this doesn't depend on the rank-based item<->Sentence alignment (both texts come from the
        # same Excel row), so it's computed for every recording, not just hit ones.
        wer, mer, cer = compute_wer_mer_cer(row["human_rater1_text"], row["human_rater2_text"])
        row["human_wer"] = wer
        row["human_mer"] = mer
        row["human_cer"] = cer

        for source, by_num in asr_by_source_ranked.items():
            item = by_num.get(n, {})
            asr_text = item.get("text")
            row[f"{source}_text"] = asr_text
            row[f"{source}_duration"] = item.get("duration")
            row[f"{source}_word_count"] = item.get("word_count")
            row[f"{source}_mean_word_prob"] = item.get("mean_word_prob")
            row[f"{source}_empty"] = item.get("empty")

            if score_quality:
                for rater in HUMAN_RATERS:
                    wer, mer, cer = compute_wer_mer_cer(row[f"human_{rater}_text"], asr_text)
                    row[f"{source}_wer_{rater}"] = wer
                    row[f"{source}_mer_{rater}"] = mer
                    row[f"{source}_cer_{rater}"] = cer

        rows.append(row)
    return rows


BASE_CSV_FIELDS = [
    "sentence_number",
    "stimulus",
    "human_rater1_text",
    "human_rater2_text",
    "human_wer",
    "human_mer",
    "human_cer",
    *[f"{s}_text" for s in ASR_SOURCES],
    *[f"{s}_{c}" for s in ASR_SOURCES for c in ("duration", "word_count", "mean_word_prob", "empty")],
]

QUALITY_CSV_FIELDS = [
    f"{s}_{m}_{rater}" for s in ASR_SOURCES for m in ("wer", "mer", "cer") for rater in HUMAN_RATERS
]


def csv_fields(score_quality: bool) -> list[str]:
    return BASE_CSV_FIELDS + QUALITY_CSV_FIELDS if score_quality else BASE_CSV_FIELDS


def main():
    human_dir = data_path("human-transcriptions")
    human_sheets = load_human_sheets(human_dir)
    log.info("Loaded %d human-transcription sheets", len(human_sheets))

    hit_list_path = data_path("transcribed_v2_hit_recordings.txt")
    hit_recordings = set(hit_list_path.read_text(encoding="utf-8").split())
    log.info("Loaded %d hit recordings for WER/MER scoring", len(hit_recordings))

    # "v2" here names this postprocessing snapshot (4 ASR sources incl. v2_filtered, +CER), not to
    # be confused with the "v2" ASR source key -- an earlier 3-source/WER+MER-only snapshot is
    # preserved at data/postprocessed/v1/ and is not regenerated by this script.
    out_root = data_path("postprocessed", "v2")
    v2_root = data_path(ASR_SOURCES["v2"][0])

    n_recordings = 0
    n_matched_human = 0
    n_scored = 0

    for version in VERSIONS:
        version_dir = v2_root / version
        if not version_dir.exists():
            continue
        for recording_dir in sorted(version_dir.iterdir()):
            if not recording_dir.is_dir():
                continue
            recording_name = recording_dir.name

            asr_by_source = {}
            for source, (folder, summary_filename) in ASR_SOURCES.items():
                src_dir = data_path(folder, version, recording_name)
                asr_by_source[source] = load_asr_items(src_dir, summary_filename)

            key = parse_folder_name(recording_name)
            human_rows = human_sheets.get(key, []) if key else []
            if key and key not in human_sheets:
                log.warning("No human transcription sheet found for %s (key=%s)", recording_name, key)
            elif human_rows:
                n_matched_human += 1

            score_quality = recording_name in hit_recordings
            rows = build_rows(human_rows, asr_by_source, score_quality)
            if score_quality:
                n_scored += 1

            out_dir = out_root / version / recording_name
            out_dir.mkdir(parents=True, exist_ok=True)
            out_path = out_dir / "transcriptions.csv"
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=csv_fields(score_quality))
                writer.writeheader()
                writer.writerows(rows)

            n_recordings += 1

    log.info(
        "Wrote %d recording CSVs (%d matched a human sheet, %d scored with WER/MER) under %s",
        n_recordings,
        n_matched_human,
        n_scored,
        out_root,
    )


if __name__ == "__main__":
    main()
